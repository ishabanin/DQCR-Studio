from __future__ import annotations

from datetime import datetime, timezone
import io
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel


TYPE_DIMENSION_PATTERN = re.compile(r"\[(\d+)(?:,\s*(\d+))?\]")


class CatalogAttribute(BaseModel):
    name: str
    display_name: str
    description: str
    domain_type: str
    raw_type: str
    is_nullable: bool
    is_key: bool
    position: int


class CatalogEntity(BaseModel):
    name: str
    display_name: str
    module: str
    info_object: str
    attributes: list[CatalogAttribute]


class CatalogMeta(BaseModel):
    source_filename: str
    loaded_at: str
    entity_count: int
    attribute_count: int
    version_label: str


class CatalogService:
    def __init__(self, catalog_path: Path):
        self.catalog_path = catalog_path
        self.catalog_file = self.catalog_path / "catalog.json"
        self.meta_file = self.catalog_path / "catalog.meta.json"

    def parse_xlsx(self, file_bytes: bytes, filename: str, version_label: str) -> tuple[list[CatalogEntity], CatalogMeta]:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            if "Атрибуты" not in workbook.sheetnames:
                raise ValueError("Sheet 'Атрибуты' not found in xlsx file.")

            sheet = workbook["Атрибуты"]
            row_iter = sheet.iter_rows(values_only=True)
            headers_row = next(row_iter, None)
            if headers_row is None:
                raise ValueError("Sheet 'Атрибуты' is empty.")

            headers = {
                str(value).strip(): idx
                for idx, value in enumerate(headers_row)
                if isinstance(value, str) and str(value).strip()
            }
            aliases = {
                "entity_display_name": ["Наименование сущности", "наименование сущности"],
                "entity_name": ["Наименование сущности (сист.)", "Объект"],
                "info_object": ["Информационный объект"],
                "module": ["Модуль", "Module"],
                "attribute_name": ["Наименование атрибута (сист.)", "атрибут"],
                "attribute_display_name": ["Наименование атрибута"],
                "attribute_description": ["Описание", "Описание атрибута"],
                "raw_type": ["Тип"],
                "nullable": ["Допустимость пустого значения"],
                "key": ["П. Н."],
            }
            header_index: dict[str, int] = {}
            missing: list[str] = []
            for field, options in aliases.items():
                idx = next((headers[item] for item in options if item in headers), None)
                if idx is None:
                    if field in {"attribute_display_name", "attribute_description"}:
                        continue
                    missing.append("/".join(options))
                else:
                    header_index[field] = idx
            if missing:
                raise ValueError(f"Missing required columns in 'Атрибуты': {', '.join(missing)}")

            entities_map: dict[str, CatalogEntity] = {}
            entity_order: list[str] = []

            for row in row_iter:
                entity_name = _to_text(_cell(row, header_index["entity_name"]))
                attr_name = _to_text(_cell(row, header_index["attribute_name"]))
                if not entity_name or not attr_name:
                    continue

                entity_display_name = _to_text(_cell(row, header_index["entity_display_name"])) or entity_name
                info_object = _to_text(_cell(row, header_index["info_object"]))
                module = _to_text(_cell(row, header_index["module"]))
                attr_display_name = _to_text(_cell(row, header_index.get("attribute_display_name", -1))) or attr_name
                attr_description = _to_text(_cell(row, header_index.get("attribute_description", -1)))
                raw_type = _to_text(_cell(row, header_index["raw_type"]))
                is_nullable = _is_nullable(_cell(row, header_index["nullable"]))
                is_key = _is_key(_cell(row, header_index["key"]))

                entity = entities_map.get(entity_name)
                if entity is None:
                    entity = CatalogEntity(
                        name=entity_name,
                        display_name=entity_display_name,
                        module=module,
                        info_object=info_object,
                        attributes=[],
                    )
                    entities_map[entity_name] = entity
                    entity_order.append(entity_name)

                entity.attributes.append(
                    CatalogAttribute(
                        name=attr_name,
                        display_name=attr_display_name,
                        description=attr_description,
                        domain_type=_normalize_domain_type(raw_type),
                        raw_type=raw_type,
                        is_nullable=is_nullable,
                        is_key=is_key,
                        position=len(entity.attributes),
                    )
                )

            entities = [entities_map[name] for name in entity_order]
            if not entities:
                raise ValueError("No valid rows found in sheet 'Атрибуты'.")

            attribute_count = sum(len(entity.attributes) for entity in entities)
            meta = CatalogMeta(
                source_filename=filename,
                loaded_at=datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
                entity_count=len(entities),
                attribute_count=attribute_count,
                version_label=version_label,
            )
            return entities, meta
        finally:
            workbook.close()

    def save(self, entities: list[CatalogEntity], meta: CatalogMeta) -> None:
        self.catalog_path.mkdir(parents=True, exist_ok=True)

        self.catalog_file.unlink(missing_ok=True)
        self.meta_file.unlink(missing_ok=True)

        self._write_json_atomic(self.catalog_file, {"entities": [entity.model_dump() for entity in entities]})
        self._write_json_atomic(self.meta_file, meta.model_dump())

    def load_meta(self) -> CatalogMeta | None:
        if not self.meta_file.exists() or not self.meta_file.is_file():
            return None
        try:
            payload = json.loads(self.meta_file.read_text(encoding="utf-8"))
            if not isinstance(payload, dict):
                return None
            return CatalogMeta.model_validate(payload)
        except (json.JSONDecodeError, OSError, ValueError):
            return None

    def is_available(self) -> bool:
        return self.catalog_file.exists() and self.catalog_file.is_file()

    def search_entities(self, query: str, limit: int = 20) -> list[dict[str, Any]]:
        matched, _ = self.search_entities_with_total(query=query, limit=limit)
        return matched

    def search_entities_with_total(self, query: str, limit: int = 20) -> tuple[list[dict[str, Any]], int]:
        normalized_query = query.strip().lower()
        matched: list[dict[str, Any]] = []
        total = 0
        for entity in self.list_entities():
            if (
                normalized_query
                and normalized_query not in entity.name.lower()
                and normalized_query not in entity.display_name.lower()
                and normalized_query not in entity.module.lower()
            ):
                continue
            total += 1
            if len(matched) < limit:
                matched.append(
                    {
                        "name": entity.name,
                        "display_name": entity.display_name,
                        "module": entity.module,
                        "attribute_count": len(entity.attributes),
                    }
                )
        return matched, total

    def get_entity(self, name: str) -> CatalogEntity | None:
        wanted = name.strip().lower()
        if not wanted:
            return None
        for entity in self.list_entities():
            if entity.name.lower() == wanted:
                return entity
        return None

    def list_entities(self) -> list[CatalogEntity]:
        if not self.catalog_file.exists() or not self.catalog_file.is_file():
            return []
        try:
            payload = json.loads(self.catalog_file.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
        entities_raw = payload.get("entities") if isinstance(payload, dict) else None
        entities_list = entities_raw if isinstance(entities_raw, list) else []
        entities: list[CatalogEntity] = []
        for item in entities_list:
            if not isinstance(item, dict):
                continue
            try:
                entities.append(CatalogEntity.model_validate(item))
            except ValueError:
                continue
        return entities

    def _write_json_atomic(self, path: Path, payload: dict[str, Any]) -> None:
        tmp_path = path.with_suffix(f"{path.suffix}.tmp")
        tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_path.replace(path)


def _cell(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _is_nullable(value: Any) -> bool:
    return _to_text(value).lower() == "да"


def _is_key(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return float(value) == 0

    normalized = _to_text(value)
    if not normalized:
        return False
    try:
        return float(normalized.replace(",", ".")) == 0
    except ValueError:
        return False


def _normalize_domain_type(raw_type: str) -> str:
    normalized = _to_text(raw_type)
    if not normalized:
        return "unknown"

    collapsed = re.sub(r"\s+", " ", normalized).strip()
    lower = collapsed.lower()

    if lower == "длинное целое число":
        return "bigint"
    if lower == "целое число":
        return "int"
    if lower == "большое дробное число":
        return "float"
    if lower == "дата и время":
        return "datetime"

    if lower.startswith("число"):
        dimensions = TYPE_DIMENSION_PATTERN.search(lower)
        if dimensions:
            precision = dimensions.group(1)
            scale = dimensions.group(2)
            if precision and scale:
                return f"decimal({precision},{scale})"
        return "decimal"

    if lower.startswith("строка"):
        dimensions = TYPE_DIMENSION_PATTERN.search(lower)
        if dimensions and dimensions.group(1):
            return f"varchar({dimensions.group(1)})"

    return "unknown"
