from fastapi.testclient import TestClient
import pytest

from app.routers import projects as projects_router
from pathlib import Path
import json
import io

from app.core.config import settings
from openpyxl import Workbook


def _build_catalog_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Атрибуты"
    sheet.append(
        [
            "Наименование сущности",
            "Наименование сущности (сист.)",
            "Информационный объект",
            "Модуль",
            "Наименование атрибута (сист.)",
            "Наименование атрибута",
            "Описание",
            "Тип",
            "Допустимость пустого значения",
            "П. Н.",
        ]
    )
    sheet.append(["Аналитический счет", "Account", "ГК", "DWH", "ID", "ИД", "Ключевой идентификатор", "Длинное целое число", "нет", 0])
    sheet.append(["Аналитический счет", "Account", "ГК", "DWH", "BranchID", "ИДСтруктураКомпании", "ИД филиала", "Число [19, 0]", "нет", 1])
    sheet.append(["Аналитический счет", "Account", "ГК", "DWH", "Title", "Наименование", "Наименование счета", "Строка [30]", "да", 1])
    sheet.append(["Клиент", "Client", "ГК", "DWH", "CreatedAt", "ДатаСоздания", "Дата создания клиента", "Дата и время", "нет", 1])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _create_external_project(root: Path, name: str) -> Path:
    project_path = root / name
    (project_path / "contexts").mkdir(parents=True, exist_ok=True)
    (project_path / "model" / "SampleModel" / "workflow" / "01_stage").mkdir(parents=True, exist_ok=True)
    (project_path / "parameters").mkdir(parents=True, exist_ok=True)
    (project_path / "project.yml").write_text("name: External\n", encoding="utf-8")
    (project_path / "contexts" / "default.yml").write_text("name: default\nenabled: true\n", encoding="utf-8")
    (project_path / "model" / "SampleModel" / "model.yml").write_text("workflow:\n  folders: {}\n", encoding="utf-8")
    return project_path


def _create_model(project_root: Path, model_id: str) -> None:
    model_root = project_root / "model" / model_id
    workflow_root = model_root / "workflow" / "01_stage"
    workflow_root.mkdir(parents=True, exist_ok=True)
    (model_root / "model.yml").write_text(
        "\n".join(
            [
                "target_table:",
                f"  name: {model_id.lower()}_table",
                "  schema: dm",
                "  attributes:",
                "    - name: id",
                "      domain_type: number",
                "workflow:",
                "  folders:",
                "    01_stage:",
                "      enabled: true",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    (workflow_root / "folder.yml").write_text("materialized: insert_fc\n", encoding="utf-8")
    (workflow_root / "001_main.sql").write_text("-- comment\nSELECT 1 AS id\n", encoding="utf-8")


def test_catalog_upload_and_status(api_client: TestClient) -> None:
    payload = _build_catalog_xlsx_bytes()
    upload_response = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("CBR_DWH71_DataModel_templete.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"version_label": "DWH 7.1"},
    )
    assert upload_response.status_code == 200, upload_response.text
    body = upload_response.json()
    assert body["available"] is True
    assert body["meta"]["entity_count"] == 2
    assert body["meta"]["attribute_count"] == 4
    assert body["meta"]["version_label"] == "DWH 7.1"

    status_response = api_client.get("/api/v1/catalog")
    assert status_response.status_code == 200
    status_payload = status_response.json()
    assert status_payload["available"] is True
    assert status_payload["meta"]["source_filename"] == "CBR_DWH71_DataModel_templete.xlsx"

    catalog_root = Path(settings.catalog_path)
    assert (catalog_root / "catalog.json").exists()
    assert (catalog_root / "catalog.meta.json").exists()


def test_catalog_upload_replaces_previous_catalog(api_client: TestClient) -> None:
    first_payload = _build_catalog_xlsx_bytes()
    first = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("first.xlsx", first_payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert first.status_code == 200, first.text

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Атрибуты"
    sheet.append(
        [
            "Наименование сущности",
            "Наименование сущности (сист.)",
            "Информационный объект",
            "Модуль",
            "Наименование атрибута (сист.)",
            "Наименование атрибута",
            "Описание",
            "Тип",
            "Допустимость пустого значения",
            "П. Н.",
        ]
    )
    sheet.append(["Контрагент", "Counterparty", "ГК", "DWH", "CounterpartyID", "ИД", "ИД контрагента", "Целое число", "нет", 0])
    replacement_payload = io.BytesIO()
    workbook.save(replacement_payload)
    workbook.close()

    second = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("second.xlsx", replacement_payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert second.status_code == 200, second.text
    assert second.json()["meta"]["entity_count"] == 1

    entities_response = api_client.get("/api/v1/catalog/entities")
    assert entities_response.status_code == 200
    entities = entities_response.json()["entities"]
    assert len(entities) == 1
    assert entities[0]["name"] == "Counterparty"


def test_catalog_upload_validation_errors(api_client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    bad_ext = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("catalog.txt", b"not an xlsx", "text/plain")},
    )
    assert bad_ext.status_code == 422

    from app.routers import catalog as catalog_router

    monkeypatch.setattr(catalog_router, "_MAX_UPLOAD_BYTES", 8)
    too_large = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("catalog.xlsx", b"0123456789", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert too_large.status_code == 413


def test_catalog_entities_search_and_get(api_client: TestClient) -> None:
    upload_response = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("catalog.xlsx", _build_catalog_xlsx_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"version_label": "DWH 7.1"},
    )
    assert upload_response.status_code == 200

    search_acc = api_client.get("/api/v1/catalog/entities", params={"search": "acc"})
    assert search_acc.status_code == 200
    assert any(item["name"] == "Account" for item in search_acc.json()["entities"])

    search_ru = api_client.get("/api/v1/catalog/entities", params={"search": "счет"})
    assert search_ru.status_code == 200
    assert any(item["name"] == "Account" for item in search_ru.json()["entities"])

    entity_response = api_client.get("/api/v1/catalog/entities/Account")
    assert entity_response.status_code == 200
    entity_payload = entity_response.json()
    assert len(entity_payload["attributes"]) == 3
    by_name = {item["name"]: item for item in entity_payload["attributes"]}
    assert by_name["ID"]["domain_type"] == "bigint"
    assert by_name["ID"]["description"] == "Ключевой идентификатор"
    assert by_name["BranchID"]["domain_type"] == "decimal(19,0)"
    assert by_name["Title"]["domain_type"] == "varchar(30)"

    not_found = api_client.get("/api/v1/catalog/entities/NonExistent")
    assert not_found.status_code == 404


def test_catalog_upload_supports_alias_headers(api_client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Атрибуты"
    sheet.append(
        [
            "наименование сущности",
            "Объект",
            "Информационный объект",
            "Module",
            "атрибут",
            "Наименование атрибута",
            "Описание атрибута",
            "Тип",
            "Допустимость пустого значения",
            "П. Н.",
        ]
    )
    sheet.append(["Карточка клиента", "ClientCard", "ГК", "CRM", "ClientCardID", "ИДКарточки", "Идентификатор карточки", "Целое число", "нет", 0])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    upload_response = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("catalog.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200, upload_response.text

    entity_response = api_client.get("/api/v1/catalog/entities/ClientCard")
    assert entity_response.status_code == 200, entity_response.text
    item = entity_response.json()
    assert item["module"] == "CRM"
    assert item["display_name"] == "Карточка клиента"
    assert item["attributes"][0]["name"] == "ClientCardID"
    assert item["attributes"][0]["description"] == "Идентификатор карточки"


def test_catalog_entities_requires_loaded_catalog(api_client: TestClient) -> None:
    response = api_client.get("/api/v1/catalog/entities")
    assert response.status_code == 404


def test_save_model_persists_attributes_and_target_table_table(api_client: TestClient) -> None:
    payload = {
        "model": {
            "target_table": {
                "name": "Account",
                "table": "Account",
                "schema": "dbo",
                "attributes": [
                    {"name": "ID", "domain_type": "bigint", "is_key": True, "required": True},
                    {"name": "BranchID", "domain_type": "decimal", "is_key": False, "required": True},
                ],
            },
            "workflow": {"description": "wf", "folders": [{"id": "01_stage", "enabled": True}]},
            "cte_settings": {"default": "insert_fc", "by_context": {}},
        }
    }
    response = api_client.put("/api/v1/projects/demo/models/SampleModel", json=payload)
    assert response.status_code == 200, response.text

    model_path = Path(settings.projects_path) / "demo" / "model" / "SampleModel" / "model.yml"
    raw = model_path.read_text(encoding="utf-8")
    assert "table: Account" in raw
    assert "fields:" not in raw
    assert "attributes:" in raw
    assert "domain_type: bigint" in raw
    assert "required: true" in raw
    assert "entity_ref" not in raw


def test_save_model_creates_workflow_root_when_missing(api_client: TestClient) -> None:
    project_root = Path(settings.projects_path) / "demo"
    model_root = project_root / "model" / "SampleModel"
    workflow_root = model_root / "workflow"
    if workflow_root.exists():
        for child in sorted(workflow_root.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            elif child.is_dir():
                child.rmdir()
        workflow_root.rmdir()

    payload = {
        "model": {
            "target_table": {
                "name": "Account",
                "table": "Account",
                "schema": "dbo",
                "attributes": [],
            },
            "workflow": {"description": "wf", "folders": [{"id": "01_stage", "enabled": True}]},
        }
    }
    response = api_client.put("/api/v1/projects/demo/models/SampleModel", json=payload)
    assert response.status_code == 200, response.text
    assert (model_root / "workflow" / "01_stage").exists()
    assert (model_root / "workflow" / "01_stage" / "folder.yml").exists()


def test_autocomplete_includes_columns_from_target_table_attributes(api_client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    model_path = Path(settings.projects_path) / "demo" / "model" / "SampleModel" / "model.yml"
    model_path.write_text(
        "\n".join(
            [
                "target_table:",
                "  name: Account",
                "  schema: dbo",
                "  attributes:",
                "    - name: ID",
                "      domain_type: bigint",
                "      is_key: true",
                "workflow:",
                "  folders:",
                "    01_stage:",
                "      enabled: true",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(projects_router, "_ensure_workflow_payload", lambda _p, _m, force_rebuild=False: None)
    response = api_client.get("/api/v1/projects/demo/autocomplete", params={"model_id": "SampleModel"})
    assert response.status_code == 200, response.text
    payload = response.json()
    objects = payload.get("objects", [])
    target_obj = next((item for item in objects if item.get("kind") == "target_table" and item.get("model_id") == "SampleModel"), None)
    assert target_obj is not None
    columns = target_obj.get("columns", [])
    assert any(str(column.get("name", "")).lower() == "id" for column in columns)


def test_autocomplete_includes_catalog_entities_when_catalog_loaded(api_client: TestClient) -> None:
    upload_response = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("catalog.xlsx", _build_catalog_xlsx_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200, upload_response.text

    response = api_client.get("/api/v1/projects/demo/autocomplete", params={"model_id": "SampleModel"})
    assert response.status_code == 200, response.text
    body = response.json()
    catalog_object = next((item for item in body["objects"] if item.get("kind") == "catalog_entity" and item.get("name") == "Account"), None)
    assert catalog_object is not None
    assert catalog_object["source"] == "catalog"
    assert catalog_object["module"] == "DWH"
    assert catalog_object["object_name"] == "Account"
    assert "_m.DWH.Account" in catalog_object["lookup_keys"]
    assert len(catalog_object["columns"]) == 3
    by_name = {item["name"]: item for item in catalog_object["columns"]}
    assert by_name["ID"]["domain_type"] == "bigint"
    assert by_name["ID"]["is_key"] is True
    assert by_name["ID"]["description"] == "Ключевой идентификатор"


def test_autocomplete_without_catalog_keeps_previous_behavior(api_client: TestClient) -> None:
    response = api_client.get("/api/v1/projects/demo/autocomplete", params={"model_id": "SampleModel"})
    assert response.status_code == 200, response.text
    body = response.json()
    assert all(item.get("kind") != "catalog_entity" for item in body["objects"])


def test_autocomplete_prefers_project_object_when_catalog_name_conflicts(api_client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    model_path = Path(settings.projects_path) / "demo" / "model" / "SampleModel" / "model.yml"
    model_path.write_text(
        "\n".join(
            [
                "target_table:",
                "  name: Account",
                "  attributes:",
                "    - name: ID",
                "      domain_type: bigint",
                "      is_key: true",
                "workflow:",
                "  folders:",
                "    01_stage:",
                "      enabled: true",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    upload_response = api_client.post(
        "/api/v1/catalog/upload",
        files={"file": ("catalog.xlsx", _build_catalog_xlsx_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200, upload_response.text

    monkeypatch.setattr(projects_router, "_ensure_workflow_payload", lambda _p, _m, force_rebuild=False: None)
    response = api_client.get("/api/v1/projects/demo/autocomplete", params={"model_id": "SampleModel"})
    assert response.status_code == 200, response.text
    body = response.json()

    account_matches = [
        item
        for item in body["objects"]
        if any(str(key).strip().lower() == "account" for key in item.get("lookup_keys", []))
    ]
    assert len(account_matches) == 1
    account_object = account_matches[0]
    assert account_object["kind"] == "target_table"
    assert account_object["source"] == "project_model_fallback"
    assert {column["name"] for column in account_object["columns"]} == {"ID", "BranchID", "Title"}


def test_get_model_migrates_legacy_fields_into_attributes(api_client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    model_path = Path(settings.projects_path) / "demo" / "model" / "SampleModel" / "model.yml"
    model_path.write_text(
        "\n".join(
            [
                "target_table:",
                "  name: Account",
                "  schema: dbo",
                "  attributes:",
                "fields:",
                "  - name: ID",
                "    type: bigint",
                "    is_key: true",
                "  - name: ID",
                "    type: bigint",
                "    is_key: true",
                "  - name: BranchID",
                "    type: decimal",
                "    is_key: false",
                "workflow:",
                "  folders:",
                "    01_stage:",
                "      enabled: true",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(projects_router, "_ensure_workflow_payload", lambda _p, _m, force_rebuild=False: None)

    response = api_client.get("/api/v1/projects/demo/models/SampleModel")
    assert response.status_code == 200, response.text
    payload = response.json()
    model = payload["model"]
    attrs = model["target_table"]["attributes"]
    assert len(attrs) == 2
    assert attrs[0]["name"] == "ID"
    assert attrs[0]["domain_type"] == "bigint"
    assert attrs[0]["is_key"] is True
    assert attrs[1]["name"] == "BranchID"
    assert "fields" not in model


def test_projects_list_and_create(api_client: TestClient) -> None:
    response = api_client.get("/api/v1/projects")
    assert response.status_code == 200, response.text
    ids = [item["id"] for item in response.json()]
    assert "demo" in ids

    create_response = api_client.post(
        "/api/v1/projects",
        json={
            "project_id": "newproj",
            "name": "New Project",
            "description": "from test",
            "template": "flx",
            "contexts": ["default", "dev"],
            "properties": {"owner": "qa"},
            "model": {
                "name": "SampleModel",
                "first_folder": "01_stage",
                "attributes": [{"name": "id", "domain_type": "number", "is_key": True}],
            },
        },
    )
    assert create_response.status_code == 200
    assert create_response.json()["id"] == "newproj"

    contexts_response = api_client.get("/api/v1/projects/newproj/contexts")
    assert contexts_response.status_code == 200
    context_ids = [item["id"] for item in contexts_response.json()]
    assert "default" in context_ids
    assert "dev" in context_ids


def test_projects_import_and_connect(api_client: TestClient, tmp_path: Path) -> None:
    external_root = tmp_path / "external"
    external_root.mkdir(parents=True, exist_ok=True)
    import_source = _create_external_project(external_root, "import_source")
    connect_source = _create_external_project(external_root, "connect_source")

    import_response = api_client.post(
        "/api/v1/projects",
        json={
            "mode": "import",
            "source_path": str(import_source),
            "project_id": "imported_proj",
            "name": "Imported Project",
        },
    )
    assert import_response.status_code == 200
    assert import_response.json()["id"] == "imported_proj"
    assert import_response.json()["source_type"] == "imported"

    connect_response = api_client.post(
        "/api/v1/projects",
        json={
            "mode": "connect",
            "source_path": str(connect_source),
            "project_id": "linked_proj",
            "name": "Linked Project",
        },
    )
    assert connect_response.status_code == 200
    assert connect_response.json()["id"] == "linked_proj"
    assert connect_response.json()["source_type"] == "linked"

    projects_response = api_client.get("/api/v1/projects")
    assert projects_response.status_code == 200
    project_rows = {item["id"]: item for item in projects_response.json()}
    assert project_rows["imported_proj"]["source_type"] == "imported"
    assert project_rows["linked_proj"]["source_type"] == "linked"
    assert project_rows["linked_proj"]["availability_status"] == "available"

    linked_contexts = api_client.get("/api/v1/projects/linked_proj/contexts")
    assert linked_contexts.status_code == 200
    assert linked_contexts.json()[0]["id"] == "default"


def test_projects_import_upload(api_client: TestClient) -> None:
    files = [
        ("files", ("project.yml", b"name: Uploaded Project\n", "text/plain")),
        ("files", ("default.yml", b"name: default\nenabled: true\n", "text/plain")),
        ("files", ("model.yml", b"workflow:\n  folders: {}\n", "text/plain")),
    ]
    data = {
        "relative_paths": ["project.yml", "contexts/default.yml", "model/SampleModel/model.yml"],
        "project_id": "uploaded_proj",
        "name": "Uploaded Project",
    }
    response = api_client.post("/api/v1/projects/import-upload", files=files, data=data)
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["id"] == "uploaded_proj"
    assert body["source_type"] == "imported"
    assert body["source_path"] is None


def test_create_model_scaffold(api_client: TestClient) -> None:
    response = api_client.post("/api/v1/projects/demo/files/model", json={"model_id": "RevenueMart"})
    assert response.status_code == 200, response.text

    body = response.json()
    assert body["model_id"] == "RevenueMart"
    assert body["path"] == "model/RevenueMart"
    assert body["file_path"] == "model/RevenueMart/model.yml"

    content_response = api_client.get("/api/v1/projects/demo/files/content", params={"path": "model/RevenueMart/model.yml"})
    assert content_response.status_code == 200
    assert content_response.json()["content"] == "target_table:\n  attributes:\n\nworkflow:\n\n  folders:\n"

    model_response = api_client.get("/api/v1/projects/demo/models/RevenueMart")
    assert model_response.status_code == 200, model_response.text
    model_body = model_response.json()
    assert model_body["model_id"] == "RevenueMart"
    assert model_body["path"] == "model/RevenueMart/model.yml"
    assert model_body["model"]["target_table"]["attributes"] == []
    assert model_body["model"]["workflow"]["folders"] == []


def test_validate_api_and_history(api_client: TestClient) -> None:
    response = api_client.post(
        "/api/v1/projects/demo/validate",
        json={"model_id": "SampleModel", "categories": ["general", "sql", "descriptions"]},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["project"] == "demo"
    assert body["model"] == "SampleModel"
    assert "summary" in body
    assert isinstance(body["rules"], list)
    assert len(body["rules"]) > 0
    assert "workflow_updated_at" in body

    history_response = api_client.get("/api/v1/projects/demo/validate/history")
    assert history_response.status_code == 200
    history = history_response.json()
    assert len(history) >= 1
    assert history[0]["run_id"] == body["run_id"]
    assert history[0]["workflow_updated_at"] == body["workflow_updated_at"]


def test_build_api_history_and_files(api_client: TestClient) -> None:
    build_response = api_client.post(
        "/api/v1/projects/demo/build",
        json={
            "model_id": "SampleModel",
            "engine": "dqcr",
            "context": "default",
            "dry_run": False,
        },
    )
    assert build_response.status_code == 200
    build = build_response.json()
    build_id = build["build_id"]
    assert build["files_count"] >= 1
    assert "workflow_updated_at" in build

    history_response = api_client.get("/api/v1/projects/demo/build/history")
    assert history_response.status_code == 200
    assert history_response.json()[0]["build_id"] == build_id
    assert history_response.json()[0]["workflow_updated_at"] == build["workflow_updated_at"]

    files_response = api_client.get(f"/api/v1/projects/demo/build/{build_id}/files")
    assert files_response.status_code == 200
    files_payload = files_response.json()
    assert files_payload["build_id"] == build_id
    assert len(files_payload["files"]) >= 1

    download_response = api_client.get(f"/api/v1/projects/demo/build/{build_id}/download")
    assert download_response.status_code == 200
    assert download_response.headers["content-type"].startswith("application/zip")


def test_build_preview_uses_engine_from_build_record(api_client: TestClient) -> None:
    build_response = api_client.post(
        "/api/v1/projects/demo/build",
        json={
            "model_id": "SampleModel",
            "engine": "airflow",
            "context": "default",
            "dry_run": False,
        },
    )
    assert build_response.status_code == 200, build_response.text
    build_id = build_response.json()["build_id"]

    preview_response = api_client.post(
        f"/api/v1/projects/demo/build/{build_id}/preview",
        json={
            "model_id": "SampleModel",
            "sql_path": "model/SampleModel/workflow/01_stage/001_main.sql",
        },
    )
    assert preview_response.status_code == 200, preview_response.text
    body = preview_response.json()
    assert body["build_id"] == build_id
    assert body["engine"] == "airflow"
    assert body["preview"].startswith("-- airflow.sql preview")


def test_build_preview_fails_for_unknown_build_id(api_client: TestClient) -> None:
    preview_response = api_client.post(
        "/api/v1/projects/demo/build/bld-does-not-exist/preview",
        json={
            "model_id": "SampleModel",
            "sql_path": "model/SampleModel/workflow/01_stage/001_main.sql",
        },
    )
    assert preview_response.status_code == 404
    assert "not found" in preview_response.json()["detail"].lower()


def test_build_history_persists_after_memory_reset(api_client: TestClient) -> None:
    build_response = api_client.post(
        "/api/v1/projects/demo/build",
        json={
            "model_id": "SampleModel",
            "engine": "dqcr",
            "context": "default",
            "dry_run": False,
        },
    )
    assert build_response.status_code == 200
    build_id = build_response.json()["build_id"]

    projects_router._BUILD_HISTORY.clear()

    history_response = api_client.get("/api/v1/projects/demo/build/history")
    assert history_response.status_code == 200
    history = history_response.json()
    assert len(history) >= 1
    assert history[0]["build_id"] == build_id


def test_build_history_discovers_existing_build_dirs_without_history_file(api_client: TestClient) -> None:
    project_root = Path(projects_router.settings.projects_path) / "demo"
    build_dir = project_root / ".dqcr_builds" / "bld-legacy1234" / "SampleModel" / "workflow" / "01_stage"
    build_dir.mkdir(parents=True, exist_ok=True)
    (build_dir / "001_main.sql").write_text("select 1;\n", encoding="utf-8")

    history_file = project_root / ".dqcr_builds" / "history.json"
    if history_file.exists():
        history_file.unlink()
    projects_router._BUILD_HISTORY.clear()

    history_response = api_client.get("/api/v1/projects/demo/build/history")
    assert history_response.status_code == 200
    history = history_response.json()
    assert len(history) >= 1
    assert history[0]["build_id"] == "bld-legacy1234"
    assert history[0]["files_count"] >= 1
    assert history[0]["model"] == "SampleModel"


def test_build_history_discovers_builds_in_custom_output_paths(api_client: TestClient) -> None:
    project_root = Path(projects_router.settings.projects_path) / "demo"
    build_dir = project_root / "custom_output" / "bld-custom1234" / "SampleModel"
    build_dir.mkdir(parents=True, exist_ok=True)
    (build_dir / "result.sql").write_text("select 42;\n", encoding="utf-8")

    history_file = project_root / ".dqcr_builds" / "history.json"
    if history_file.exists():
        history_file.unlink()
    projects_router._BUILD_HISTORY.clear()

    history_response = api_client.get("/api/v1/projects/demo/build/history")
    assert history_response.status_code == 200
    history = history_response.json()
    custom_item = next((item for item in history if item["build_id"] == "bld-custom1234"), None)
    assert custom_item is not None
    assert custom_item["output_path"].startswith("custom_output/bld-custom1234")
    assert custom_item["files_count"] >= 1
def test_workflow_status_and_model_workflow_endpoints(api_client: TestClient) -> None:
    status_response = api_client.get("/api/v1/projects/demo/workflow/status")
    assert status_response.status_code == 200
    status_payload = status_response.json()
    assert status_payload["project_id"] == "demo"
    assert isinstance(status_payload["models"], list)

    model_response = api_client.get("/api/v1/projects/demo/models/SampleModel/workflow")
    assert model_response.status_code == 200
    model_payload = model_response.json()
    assert model_payload["project_id"] == "demo"
    assert model_payload["model_id"] == "SampleModel"
    assert model_payload["status"] in {"ready", "stale", "building", "error", "missing"}

    rebuild_response = api_client.post("/api/v1/projects/demo/models/SampleModel/workflow/rebuild")
    assert rebuild_response.status_code == 200
    rebuild_payload = rebuild_response.json()
    assert rebuild_payload["model_id"] == "SampleModel"
    assert rebuild_payload["status"] in {"ready", "stale", "building", "error", "missing"}


def test_files_create_folder_and_content(api_client: TestClient) -> None:
    folder_response = api_client.post(
        "/api/v1/projects/demo/files/folder",
        json={"path": "model/SampleModel/workflow/02_new"},
    )
    assert folder_response.status_code == 200
    assert folder_response.json()["path"] == "model/SampleModel/workflow/02_new"

    file_response = api_client.put(
        "/api/v1/projects/demo/files/content",
        json={
            "path": "model/SampleModel/workflow/02_new/001_extra.sql",
            "content": "select 1\n",
        },
    )
    assert file_response.status_code == 200

    read_response = api_client.get(
        "/api/v1/projects/demo/files/content",
        params={"path": "model/SampleModel/workflow/02_new/001_extra.sql"},
    )
    assert read_response.status_code == 200
    assert read_response.json()["content"] == "select 1\n"


def test_lineage_uses_workflow_order_for_edges(api_client: TestClient) -> None:
    reorder_response = api_client.put(
        "/api/v1/projects/demo/models/SampleModel",
        json={
            "model": {
                "target_table": {
                    "name": "sample_table",
                    "schema": "dm",
                    "description": "sample",
                    "template": "dqcr",
                    "engine": "oracle",
                    "attributes": [{"name": "id", "domain_type": "number"}],
                },
                "workflow": {
                    "description": "sample workflow",
                    "folders": [
                        {"id": "02_publish", "description": "publish", "enabled": True},
                        {"id": "01_stage", "description": "stage", "enabled": True},
                    ],
                },
                "cte_settings": {"default": "insert_fc", "by_context": {}},
            }
        },
    )
    assert reorder_response.status_code == 200

    folder_response = api_client.post(
        "/api/v1/projects/demo/files/folder",
        json={"path": "model/SampleModel/workflow/02_publish"},
    )
    assert folder_response.status_code == 200

    file_response = api_client.put(
        "/api/v1/projects/demo/files/content",
        json={
            "path": "model/SampleModel/workflow/02_publish/001_publish.sql",
            "content": "select * from stage\n",
        },
    )
    assert file_response.status_code == 200

    lineage_response = api_client.get("/api/v1/projects/demo/models/SampleModel/lineage")
    assert lineage_response.status_code == 200
    lineage = lineage_response.json()

    assert [node["id"] for node in lineage["nodes"][:2]] == ["02_publish", "01_stage"]
    assert lineage["edges"][0]["source"] == "02_publish"
    assert lineage["edges"][0]["target"] == "01_stage"


def test_lineage_prefers_framework_workflow_cache(api_client: TestClient, tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    publish_dir = projects_root / "demo" / "model" / "SampleModel" / "workflow" / "02_publish"
    publish_dir.mkdir(parents=True, exist_ok=True)
    (publish_dir / "001_publish.sql").write_text("select 1\n", encoding="utf-8")

    cache_file = projects_root / "demo" / ".dqcr_workflow_cache" / "SampleModel.json"
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(
        json.dumps(
            {
                "steps": [
                    {
                        "full_name": "01_stage/001_main/sql",
                        "step_type": "sql",
                        "folder": "01_stage",
                        "dependencies": [],
                        "context": "all",
                        "sql_model": {
                            "name": "001_main",
                            "path": "model/SampleModel/workflow/01_stage/001_main.sql",
                            "materialization": "insert_fc",
                            "metadata": {"parameters": ["date_start"], "cte": {"stg": {}}},
                        },
                    },
                    {
                        "full_name": "02_publish/001_publish/sql",
                        "step_type": "sql",
                        "folder": "02_publish",
                        "dependencies": ["01_stage/001_main/sql"],
                        "context": "default",
                        "sql_model": {
                            "name": "001_publish",
                            "path": "model/SampleModel/workflow/02_publish/001_publish.sql",
                            "materialization": "upsert_fc",
                            "metadata": {"parameters": ["date_end"], "cte": {}},
                        },
                    },
                    {
                        "full_name": "02_publish/sql_virtual/cte/tmp",
                        "step_type": "sql",
                        "folder": "02_publish",
                        "dependencies": ["02_publish/001_publish/sql"],
                        "context": "default",
                        "sql_model": {
                            "name": "sql_virtual_tmp",
                            "path": ".",
                            "materialization": "ephemeral",
                            "metadata": {"parameters": ["ghost_param"], "cte": {"ghost_cte": {}}},
                        },
                    },
                ],
                "config": {
                    "folders": {
                        "01_stage": {"materialized": "insert_fc"},
                        "02_publish": {"materialized": "upsert_fc"},
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    lineage_response = api_client.get("/api/v1/projects/demo/models/SampleModel/lineage")
    assert lineage_response.status_code == 200
    lineage = lineage_response.json()
    assert [node["id"] for node in lineage["nodes"]] == ["01_stage", "02_publish"]
    assert lineage["edges"] == [
        {
            "id": "01_stage->02_publish",
            "source": "01_stage",
            "target": "02_publish",
            "status": "resolved",
        }
    ]
    stage_node = next(node for node in lineage["nodes"] if node["id"] == "01_stage")
    publish_node = next(node for node in lineage["nodes"] if node["id"] == "02_publish")
    assert stage_node["queries"] == ["001_main.sql"]
    assert publish_node["queries"] == ["001_publish.sql"]
    assert publish_node["ctes"] == []
    assert lineage["summary"]["params"] == 2


def test_lineage_builds_with_requested_context(api_client: TestClient, monkeypatch) -> None:
    requested_contexts: list[str | None] = []

    def _fake_run_workflow_build(project_id: str, model_id: str, context: str | None = None) -> dict[str, object]:
        requested_contexts.append(context)
        return {
            "workflow": {
                "steps": [
                    {
                        "full_name": "01_stage/001_main/sql",
                        "step_type": "sql",
                        "folder": "01_stage",
                        "dependencies": [],
                        "context": context or "all",
                        "sql_model": {
                            "name": "001_main",
                            "path": "model/SampleModel/workflow/01_stage/001_main.sql",
                            "materialization": "insert_fc",
                            "metadata": {"parameters": [], "cte": {}},
                        },
                    }
                ],
                "config": {"folders": {"01_stage": {"materialized": "insert_fc"}}},
            }
        }

    monkeypatch.setattr(projects_router.FW_SERVICE, "run_workflow_build", _fake_run_workflow_build)

    lineage_response = api_client.get("/api/v1/projects/demo/models/SampleModel/lineage", params={"context": "default"})
    assert lineage_response.status_code == 200
    lineage = lineage_response.json()

    assert requested_contexts == ["default"]
    assert [node["id"] for node in lineage["nodes"]] == ["01_stage"]
    assert lineage["nodes"][0]["queries"] == ["001_main.sql"]


def test_model_object_prefers_framework_workflow_cache(api_client: TestClient, tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    cache_file = projects_root / "demo" / ".dqcr_workflow_cache" / "SampleModel.json"
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(
        json.dumps(
            {
                "target_table": {
                    "name": "sample_table_from_build",
                    "schema": "dm_build",
                    "description": "from workflow cache",
                    "attributes": [
                        {"name": "id", "domain_type": "number", "is_key": True},
                        {"name": "amount", "domain_type": "number", "required": True},
                    ],
                },
                "config": {
                    "description": "workflow from build",
                    "folders": {
                        "01_stage": {"materialized": "insert_fc", "enabled": {}},
                        "02_publish": {"materialized": "upsert_fc", "enabled": {"contexts": ["default"]}},
                    },
                    "cte": {"cte_materialization": "stage_calcid", "by_context": {"default": "insert_fc"}},
                },
                "steps": [
                    {
                        "full_name": "01_stage/001_main/sql",
                        "step_type": "sql",
                        "folder": "01_stage",
                        "dependencies": [],
                        "context": "all",
                        "sql_model": {"name": "001_main", "path": "model/SampleModel/workflow/01_stage/001_main.sql"},
                    },
                    {
                        "full_name": "02_publish/001_publish/sql",
                        "step_type": "sql",
                        "folder": "02_publish",
                        "dependencies": ["01_stage/001_main/sql"],
                        "context": "default",
                        "sql_model": {"name": "001_publish", "path": "model/SampleModel/workflow/02_publish/001_publish.sql"},
                    },
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    response = api_client.get("/api/v1/projects/demo/models/SampleModel")
    assert response.status_code == 200
    body = response.json()
    assert body["model"]["target_table"]["name"] == "sample_table_from_build"
    assert body["model"]["workflow"]["description"] == "workflow from build"
    assert [item["id"] for item in body["model"]["workflow"]["folders"]] == ["01_stage", "02_publish"]
    assert body["model"]["cte_settings"]["default"] == "stage_calcid"


def test_project_parameters_merge_framework_workflow_cache(api_client: TestClient, tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    cache_file = projects_root / "demo" / ".dqcr_workflow_cache" / "SampleModel.json"
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(
        json.dumps(
            {
                "steps": [
                    {
                        "step_type": "param",
                        "context": "all",
                        "param_model": {
                            "name": "date_start",
                            "domain_type": "date",
                            "description": "start date from workflow",
                            "values": {
                                "all": {"type": "static", "value": "2026-01-01"},
                            },
                        },
                    },
                    {
                        "step_type": "param",
                        "context": "default",
                        "param_model": {
                            "name": "calc_mode",
                            "domain_type": "string",
                            "description": "mode from workflow",
                            "values": {
                                "default": {"type": "dynamic", "value": "select 'FULL'"},
                            },
                        },
                    },
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    response = api_client.get("/api/v1/projects/demo/parameters")
    assert response.status_code == 200
    params = response.json()
    names = {item["name"] for item in params}
    assert "date_start" in names
    assert "calc_mode" in names

    calc_mode = next(item for item in params if item["name"] == "calc_mode")
    assert calc_mode["scope"] == "model:SampleModel"
    assert calc_mode["value_type"] == "dynamic"
    assert calc_mode["values"]["default"]["value"] == "select 'FULL'"


def test_config_chain_prefers_workflow_cache_metadata(api_client: TestClient, tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    cache_file = projects_root / "demo" / ".dqcr_workflow_cache" / "SampleModel.json"
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(
        json.dumps(
            {
                "config": {
                    "engine": "dqcr",
                    "folders": {
                        "01_stage": {"materialized": "insert_fc"},
                    },
                    "cte": {
                        "cte_materialization": "stage_calcid",
                        "by_context": {"default": "insert_fc"},
                    },
                },
                "steps": [
                    {
                        "full_name": "01_stage/001_main/sql",
                        "step_type": "sql",
                        "folder": "01_stage",
                        "context": "default",
                        "sql_model": {
                            "name": "001_main",
                            "path": "model/SampleModel/workflow/01_stage/001_main.sql",
                            "metadata": {
                                "parameters": ["date_start", "calc_mode"],
                                "cte": {"src": {}},
                                "inline_cte_configs": {"src": "append"},
                            },
                        },
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    response = api_client.get(
        "/api/v1/projects/demo/models/SampleModel/config-chain",
        params={"sql_path": "model/SampleModel/workflow/01_stage/001_main.sql"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["data_source"] == "workflow"
    assert body["fallback"] is False
    assert body["sql_metadata"]["parameters"] == ["calc_mode", "date_start"]
    assert body["sql_metadata"]["ctes"] == ["src"]
    assert body["cte_settings"]["default"] == "stage_calcid"
    resolved_engine = next(item for item in body["resolved"] if item["key"] == "engine")
    assert resolved_engine["value"] == "dqcr"


def test_autocomplete_uses_workflow_contexts_and_parameters(api_client: TestClient, tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    cache_file = projects_root / "demo" / ".dqcr_workflow_cache" / "SampleModel.json"
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(
        json.dumps(
            {
                "all_contexts": ["default", "vtb"],
                "target_table": {
                    "name": "sample_table",
                    "schema": "dm",
                    "attributes": [
                        {"name": "id", "domain_type": "number", "is_key": True},
                        {"name": "amount", "domain_type": "number", "is_key": False},
                    ],
                },
                "steps": [
                    {
                        "step_type": "param",
                        "context": "all",
                        "param_model": {
                            "name": "date_start",
                            "domain_type": "date",
                            "values": {"all": {"type": "static", "value": "2026-01-01"}},
                        },
                    },
                    {
                        "step_type": "sql",
                        "folder": "01_stage",
                        "full_name": "01_stage/001_main/sql",
                        "name": "001_main",
                        "sql_model": {
                            "name": "001_main",
                            "path": "/app/projects/demo/model/SampleModel/workflow/01_stage/001_main.sql",
                            "attributes": [
                                {"name": "id", "domain_type": "number", "is_key": True},
                                {"name": "amount", "domain_type": "number", "is_key": False},
                            ],
                            "metadata": {
                                "aliases": [{"alias": "id"}, {"alias": "amount"}],
                            },
                        },
                    },
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    response = api_client.get("/api/v1/projects/demo/autocomplete", params={"model_id": "SampleModel"})
    assert response.status_code == 200
    body = response.json()
    assert body["data_source"] == "workflow"
    assert body["fallback"] is False
    assert "vtb" in body["all_contexts"]
    param_names = {item["name"] for item in body["parameters"]}
    assert "date_start" in param_names
    object_names = {item["name"] for item in body["objects"]}
    assert "dm.sample_table" in object_names
    assert "_w.01_stage.001_main" in object_names
    target_table = next(item for item in body["objects"] if item["name"] == "dm.sample_table")
    assert {column["name"] for column in target_table["columns"]} == {"id", "amount"}
    workflow_query = next(item for item in body["objects"] if item["name"] == "_w.01_stage.001_main")
    assert workflow_query["path"] == "model/SampleModel/workflow/01_stage/001_main.sql"
    assert {column["name"] for column in workflow_query["columns"]} == {"id", "amount"}


def test_autocomplete_falls_back_to_model_yml_objects_when_workflow_missing(api_client: TestClient, monkeypatch) -> None:
    def _raise_build(*_args, **_kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(projects_router.FW_SERVICE, "run_workflow_build", _raise_build)

    response = api_client.get("/api/v1/projects/demo/autocomplete", params={"model_id": "SampleModel"})
    assert response.status_code == 200
    body = response.json()
    assert body["fallback"] is True
    assert body["data_source"] == "fallback"
    object_names = {item["name"] for item in body["objects"]}
    assert "dm.sample_table" in object_names


def test_workflow_status_missing_when_cache_absent(api_client: TestClient) -> None:
    status_response = api_client.get("/api/v1/projects/demo/workflow/status")
    assert status_response.status_code == 200
    payload = status_response.json()
    assert payload["status"] == "missing"
    sample_model = next(item for item in payload["models"] if item["model_id"] == "SampleModel")
    assert sample_model["status"] == "missing"
    assert sample_model["source"] == "framework_cli"


def test_workflow_soft_fail_marks_stale_and_fallback(api_client: TestClient, monkeypatch) -> None:
    project_root = Path(projects_router.settings.projects_path) / "demo"
    _create_model(project_root, "SampleModel")

    monkeypatch.setattr(
        projects_router.FW_SERVICE,
        "run_workflow_build",
        lambda project_id, model_id, context=None: {"workflow": {"steps": [], "config": {}}},
    )
    first_rebuild = api_client.post("/api/v1/projects/demo/models/SampleModel/workflow/rebuild")
    assert first_rebuild.status_code == 200
    assert first_rebuild.json()["status"] == "ready"

    def _raise_build(*_args, **_kwargs):
        raise RuntimeError("simulated build failure")

    monkeypatch.setattr(projects_router.FW_SERVICE, "run_workflow_build", _raise_build)

    second_rebuild = api_client.post("/api/v1/projects/demo/models/SampleModel/workflow/rebuild")
    assert second_rebuild.status_code == 200
    body = second_rebuild.json()
    assert body["status"] == "stale"
    assert body["source"] == "fallback"
    assert "simulated build failure" in (body["error"] or "")
    assert isinstance(body["workflow"], dict)

    status_response = api_client.get("/api/v1/projects/demo/workflow/status")
    assert status_response.status_code == 200
    assert status_response.json()["status"] == "stale"


def test_config_chain_fallback_when_workflow_unavailable(api_client: TestClient, monkeypatch) -> None:
    def _raise_build(*_args, **_kwargs):
        raise RuntimeError("workflow unavailable")

    monkeypatch.setattr(projects_router.FW_SERVICE, "run_workflow_build", _raise_build)
    response = api_client.get(
        "/api/v1/projects/demo/models/SampleModel/config-chain",
        params={"sql_path": "model/SampleModel/workflow/01_stage/001_main.sql"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["data_source"] == "fallback"
    assert body["fallback"] is True


def test_multi_model_partial_rebuild_updates_only_changed_model(api_client: TestClient, monkeypatch) -> None:
    project_root = Path(projects_router.settings.projects_path) / "demo"
    _create_model(project_root, "SecondModel")

    monkeypatch.setattr(
        projects_router.FW_SERVICE,
        "run_workflow_build",
        lambda project_id, model_id, context=None: {
            "workflow": {
                "steps": [
                    {
                        "step_type": "sql",
                        "folder": "01_stage",
                        "context": "all",
                        "sql_model": {"name": "001_main", "path": f"model/{model_id}/workflow/01_stage/001_main.sql"},
                    }
                ],
                "config": {},
            }
        },
    )

    first_model = api_client.post("/api/v1/projects/demo/models/SampleModel/workflow/rebuild")
    second_model = api_client.post("/api/v1/projects/demo/models/SecondModel/workflow/rebuild")
    assert first_model.status_code == 200
    assert second_model.status_code == 200

    before_status = api_client.get("/api/v1/projects/demo/workflow/status").json()
    before_sample = next(item for item in before_status["models"] if item["model_id"] == "SampleModel")
    before_second = next(item for item in before_status["models"] if item["model_id"] == "SecondModel")

    edit_response = api_client.put(
        "/api/v1/projects/demo/files/content",
        json={
            "path": "model/SampleModel/workflow/01_stage/001_main.sql",
            "content": "-- changed\nSELECT 2 AS id\n",
        },
    )
    assert edit_response.status_code == 200

    after_status = api_client.get("/api/v1/projects/demo/workflow/status").json()
    after_sample = next(item for item in after_status["models"] if item["model_id"] == "SampleModel")
    after_second = next(item for item in after_status["models"] if item["model_id"] == "SecondModel")

    assert after_sample["updated_at"] != before_sample["updated_at"]
    assert after_second["updated_at"] == before_second["updated_at"]
